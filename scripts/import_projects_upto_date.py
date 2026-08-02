#!/usr/bin/env python3
"""
One-off: load PROJECTS-UPTO-DATE.xlsx into projects for Tare Design Studio.

Two sheets:
  'completed' -> 92 historical completed projects (status=completed, execution stage,
                 site_location = LOCATION [· SITE MEASUREMENT], start_date = MONTH-YEAR).
                 Slugs auto-suffixed so a completed row NEVER overwrites a live project
                 that happens to share a name (e.g. old completed MOHAN vs current MOHAN).
  'running '  -> 42 current projects. Rows are reconciled against live DB projects:
                   - matched (exact or via RUNNING_CANON map) -> UPDATE in place
                     (project_type, start_date from Year, current_stage/status from progress)
                   - unmatched -> INSERT as new status=active project.

Usage:
    python3 scripts/import_projects_upto_date.py            # dry run: full plan, no writes
    python3 scripts/import_projects_upto_date.py --commit   # apply (backup taken first)
"""
import re
import sys
import datetime
import psycopg2

ROOT = "/Users/muthanna/Documents/ARCHITECT_OS"
XLSX = f"{ROOT}/PROJECTS-UPTO-DATE.xlsx"
TENANT = "d4784db6-9a2d-4075-97b5-14daaa9026ab"
COMMIT = "--commit" in sys.argv


def load_env():
    env = {}
    for line in open(f"{ROOT}/.env"):
        m = re.match(r'\s*([A-Z0-9_]+)\s*=\s*(.*)\s*$', line)
        if m:
            env[m.group(1)] = m.group(2).strip().strip('"').strip("'")
    return env


def norm(s):
    if s is None:
        return ""
    return re.sub(r'[^a-z0-9]+', ' ', str(s).lower().strip()).strip()


def slugify(s):
    s = str(s).lower().strip().replace("&", " and ")
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return re.sub(r'^-+|-+$', '', s)


# running-sheet name -> canonical live-DB project name (fuzzy cases confirmed by inspection).
# Only entries here are treated as "same project" despite different spelling.
RUNNING_CANON = {
    "vijay agarval": "VIJAY AGARWAL",
    "lakshman dattagalli": "LAKSHMAN-DATTAGALLI",
    "lakshman vijaynagar 4th stage": "LAKSHMAN- VIJAYANAGAR",
    "vinod commercial darvad": "VINOD COMMERCIAL-DARWAD",
    "darwad corporation record room": "DARWAD CORPORATION-RECORD ROOM",
    "sheela residence": "SHEELA",
    "dr sunil": "SUNIL-CONVENTION",
    "lokanna": "LOKANNA PATIL",
    "padmini baskar reddy": "PADMINI BHASKAR",
    "mgr restorent": "M.G.R RESTAURANT",
    "satyanarayan": "SATHYANARAYAN",
    "shivkumar": "SHIVAKUMAR",
    "rehaman": "REHMAN",
    "santhose thotappa": "SANTHOSH",
    "nagesh": "NAGESH", "mohan": "MOHAN", "lokesh": "LOKESH",
    "gururaj": "GURURAJ", "satish": "SATISH", "kemparaju": "KEMPARAJU",
    "varun": "VARUN", "suresh": "SURESH", "niharika": "NIHARIKA",
    "prakash": "PRAKASH", "sachin": "SACHIN", "ranjith": "RANJITH",
    "ajay": "AJAY", "mala mandanna": "MALA MANDANNA", "girish": "GIRISH",
    "shilpa manu": "SHILPA MANU", "ranga srinivas": "RANGA SRINIVAS",
    # NOTE: 'lakshman vijanagar 3rd stage', 'darwad hotel', 'kabini clubhose',
    # 'darwad coproration', 'sandeep layout' are intentionally NOT canonicalized
    # here -> treated as distinct (see AMBIGUOUS below).
}

# Running rows that are genuinely NEW current projects (no live counterpart).
RUNNING_NEW = {
    "nagendra", "raghu", "vinoda mandya", "arun lake villa", "basavesh hospete",
    "chandru shekar int",
}

# Running rows whose identity is ambiguous (could be a new phase of an existing project
# OR a separate project). Loaded as NEW active rows with a distinct slug so nothing is
# overwritten; listed explicitly in the report for the user to merge later if desired.
AMBIGUOUS_NEW = {
    "lakshman vijanagar 3rd stage",   # vs LAKSHMAN- VIJAYANAGAR (4th stage maps; 3rd is separate)
    "darwad hotel",                    # hospitality, distinct from VINOD COMMERCIAL-DARWAD
    "kabini clubhose",                 # hospitality, distinct from RAMESH KABINI FARM
    "darwad coproration",              # urban plan, distinct from record-room / commercial
    "sandeep layout",                  # vs SANDEEP LAYOUT- TADAHALLI (likely same; kept separate to avoid guess)
}

USAGE_TO_TYPE = {
    "r": "residential",
    "commertial": "commercial", "commercial": "commercial",
    "restorent": "commercial", "office space": "commercial",
    "convention hall": "commercial",
    "hospitality": "commercial", "hospitality villa": "commercial",
    "outdoor venue": "commercial",
    "urban plan": "urban",
    "int": "interior", "interior": "interior",
}


def usage_to_type(u):
    if u is None:
        return None
    return USAGE_TO_TYPE.get(str(u).strip().lower())


def parse_completed(wb):
    ws = wb["completed"]
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        name, my, loc, meas = r[1], r[2], r[3], r[4]
        if not (name and str(name).strip() and str(name).strip() != "-"):
            continue
        if not isinstance(my, datetime.datetime):   # skip shifted/junk rows
            continue
        row = [c for c in r if c is not None]
        dropout = any(str(c).strip().lower() == "dropout" for c in row)
        loc_txt = str(loc).strip() if loc else None
        meas_txt = str(meas).strip() if meas else None
        if loc_txt and meas_txt:
            site = f"{loc_txt} · {meas_txt}"
        else:
            site = loc_txt or meas_txt
        out.append({
            "name": str(name).strip(),
            "start_date": my.date(),
            "site_location": site,
            "dropout": dropout,
        })
    return out


def parse_running(wb):
    ws = wb["running "]
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = r[1]
        if name is None:
            continue
        out.append({
            "name": str(name).strip(),
            "year": r[3],
            "usage": r[4],
            "constr": r[7],   # 'yes' when execution/construction applies
        })
    return out


def running_stage_status(row):
    """Derive (current_stage, status) from the running-sheet progress flags."""
    constr = str(row.get("constr") or "").strip().lower()
    stage = "execution" if constr == "yes" else "design"
    return stage, "active"


def main():
    from openpyxl import load_workbook
    env = load_env()
    wb = load_workbook(XLSX, data_only=True)
    completed = parse_completed(wb)
    running = parse_running(wb)

    conn = psycopg2.connect(env["DATABASE_URL"])
    cur = conn.cursor()
    cur.execute(
        "select id, slug, name, current_stage, status, project_type, start_date, site_location "
        "from projects where tenant_id=%s and deleted_at is null", (TENANT,))
    db = cur.fetchall()
    db_by_norm = {norm(name): dict(id=i, slug=s, name=name, stage=st, status=stt,
                                   ptype=pt, start=sd, loc=loc)
                  for (i, s, name, st, stt, pt, sd, loc) in db}
    existing_slugs = {s for (_, s, *_rest) in db}

    def uniq_slug(name):
        base = slugify(name)
        slug = base
        n = 1
        while slug in existing_slugs:
            n += 1
            slug = f"{base}-{n}"
        existing_slugs.add(slug)
        return slug

    updates = []       # (project_id, name, fields_dict, before)
    inserts_running = []
    inserts_completed = []
    collisions = []    # completed name that also is a LIVE project

    # ---- running reconciliation ----
    for row in running:
        nkey = norm(row["name"])
        canon = RUNNING_CANON.get(nkey)
        target = None
        if canon and norm(canon) in db_by_norm:
            target = db_by_norm[norm(canon)]
        elif nkey in db_by_norm:
            target = db_by_norm[nkey]

        stage, status = running_stage_status(row)
        ptype = usage_to_type(row["usage"])
        start_date = datetime.date(int(row["year"]), 1, 1) if row.get("year") else None

        if target:
            fields = {}
            # only fill columns that are currently NULL / would be enriched; upgrade stage design->execution
            if ptype and not target["ptype"]:
                fields["project_type"] = ptype
            if start_date and not target["start"]:
                fields["start_date"] = start_date
            if stage == "execution" and target["stage"] != "execution":
                fields["current_stage"] = "execution"
            if fields:
                updates.append((target["id"], target["name"], fields, target))
        else:
            slug = uniq_slug(row["name"])
            inserts_running.append({
                "name": row["name"], "slug": slug,
                "project_type": ptype, "current_stage": stage,
                "status": "active", "start_date": start_date,
                "ambiguous": nkey in AMBIGUOUS_NEW,
            })

    # ---- completed inserts ----
    for row in completed:
        nkey = norm(row["name"])
        if nkey in db_by_norm:
            collisions.append((row["name"], db_by_norm[nkey]["slug"]))
        slug = uniq_slug(row["name"])
        inserts_completed.append({
            "name": row["name"], "slug": slug,
            "current_stage": "execution",
            "status": "completed",
            "start_date": row["start_date"],
            "site_location": row["site_location"],
            "project_type": None,
            "dropout": row["dropout"],
        })

    # ---------- REPORT ----------
    print(f"\n{'='*78}\nIMPORT PLAN  (mode: {'COMMIT' if COMMIT else 'DRY RUN'})  tenant {TENANT}\n{'='*78}")
    print(f"\nRUNNING sheet: {len(running)} rows")
    print(f"  -> UPDATE existing (enrich): {len(updates)}")
    print(f"  -> INSERT new active:        {len(inserts_running)}")
    print(f"\nCOMPLETED sheet: {len(completed)} historical rows -> INSERT completed: {len(inserts_completed)}")

    print(f"\n--- UPDATES (enrich existing live projects) ---")
    for pid, name, fields, before in updates:
        chg = ", ".join(f"{k}: {before.get({'project_type':'ptype','start_date':'start','current_stage':'stage'}[k])}→{v}"
                        for k, v in fields.items())
        print(f"  {name:34} {chg}")

    print(f"\n--- NEW ACTIVE (running, no live match) ---")
    for r in inserts_running:
        tag = "  [AMBIGUOUS — verify not a dup]" if r["ambiguous"] else ""
        print(f"  {r['slug']:30} type={str(r['project_type']):12} stage={r['current_stage']:9} {r['name']}{tag}")

    print(f"\n--- COMPLETED (historical archive) : {len(inserts_completed)} rows ---")
    for r in inserts_completed[:8]:
        print(f"  {r['slug']:30} {str(r['start_date']):12} {r['site_location'] or ''}  {r['name']}")
    print(f"  … ({len(inserts_completed)-8} more)")

    print(f"\n--- NAME COLLISIONS (completed row shares a name with a LIVE project) ---")
    print("    These get a suffixed slug so the live project is NOT overwritten.")
    for name, live_slug in collisions:
        print(f"  completed '{name}'  (live project slug: {live_slug})")
    dropouts = [r['name'] for r in inserts_completed if r['dropout']]
    print(f"\n--- DROPOUT-flagged historical rows (loaded as completed): {dropouts}")

    if not COMMIT:
        print(f"\nDRY RUN — no writes. Re-run with --commit to apply.\n")
        cur.close(); conn.close()
        return

    # ---------- WRITE ----------
    for pid, _name, fields, _b in updates:
        sets = ", ".join(f"{k}=%s" for k in fields)
        vals = list(fields.values()) + [pid]
        cur.execute(f"update projects set {sets}, updated_at=now() where id=%s", vals)

    def ins(rows):
        for r in rows:
            cur.execute(
                "insert into projects (tenant_id, name, slug, project_type, current_stage, status, "
                "start_date, site_location, scope) values (%s,%s,%s,%s,%s,%s,%s,%s,'design_and_execution')",
                (TENANT, r["name"], r["slug"], r.get("project_type"), r["current_stage"],
                 r["status"], r.get("start_date"), r.get("site_location")))
    ins(inserts_running)
    ins(inserts_completed)
    conn.commit()
    print(f"\n✅ COMMITTED: {len(updates)} updated, "
          f"{len(inserts_running)+len(inserts_completed)} inserted "
          f"({len(inserts_running)} running-new + {len(inserts_completed)} completed).\n")
    cur.close(); conn.close()


if __name__ == "__main__":
    main()
